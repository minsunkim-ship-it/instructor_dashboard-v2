"""v2: 회사 partial 매칭 + 일정 정확 매칭 + ambiguous 처리"""
import openpyxl, json, io, sys, re, os, urllib.parse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

xlsx_path = r"C:\Users\Day1_김민선\Downloads\instructor_db\contract.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
all_contracts = []
for sname, name_col, sched_col, loc_col, link_col in [
    ("강사_일반계약요청 (사용XXXXX)", 9, 13, 14, 7),
    ("강사_일반계약요청의 사본(달리수정중)", 10, 15, 16, 8),
    ("강사_일반계약요청_삼성전자", 8, 12, 13, 6),
    # v24-30: 변경계약 sheets 추가
    ("강사_변경계약요청", 9, 13, 14, 7),
    ("강사_변경계약요청의 (달리수정중)", 10, 15, 16, 8),
]:
    ws = wb[sname]
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        sched = ws.cell(row=r, column=sched_col).value
        loc = ws.cell(row=r, column=loc_col).value
        link = ws.cell(row=r, column=link_col).value
        if name and isinstance(name, str) and 2 <= len(name) <= 8:
            all_contracts.append({"name": name.strip(), "schedule": str(sched or ""), "location": str(loc or ""), "link": str(link or ""), "sheet": sname, "row": r})

all_rows = []
for off in [0,100,200,300]:
    with open(rf"C:\Users\Day1_김민선\Downloads\instructor_db\p{off}.json", "rb") as f:
        all_rows.extend(json.loads(f.read()).get("rows", []))

PUBLIC_KW = ["대학교","학교","진흥원","회의소","공단","연구원","공사","수력원자력","교육원","벤처캐피탈","과기대","과학기술원"]
def is_public(co):
    return co and any(k in co for k in PUBLIC_KW)

target = [r for r in all_rows if not is_public(r.get("company")) and r.get("company") and r.get("responseDate")]
print(f"분석 대상: {len(target)}")

def date_in(target_dt, sched_str, window=0):
    """target_dt ±window일 안에 schedule 일정 있는지"""
    if not target_dt or not sched_str: return False
    from datetime import datetime, timedelta
    try:
        tdt = datetime.strptime(target_dt, "%Y-%m-%d")
        # window 안 모든 날짜 시도
        for delta in range(-window, window + 1):
            chk = tdt + timedelta(days=delta)
            y, m, d = chk.year, chk.month, chk.day
            for p in [f"{y}[.\\-/년 ]+{m:02d}[.\\-/월 ]+{d:02d}", f"{y}[.\\-/년 ]+{m}[.\\-/월 ]+{d}", f"{m}\\.\\s*{d}", f"{m}/{d}"]:
                if re.search(p, sched_str): return True
    except: pass
    return False

def co_match(co, c):
    """회사명 매칭 — strict only. partial는 너무 wide (KB/S/KT 잡힘)."""
    co_clean = co.replace(" ","").replace("-","").lower()
    haystack = (c["location"] + " " + c["schedule"] + " " + c["link"]).lower().replace(" ","")
    if not co_clean or len(co_clean) < 3: return False, ""
    if co_clean in haystack:
        return True, "strict"
    # partial: 회사명 ≥4자 core word만
    core_words = re.findall(r"[가-힣]{4,}|[A-Za-z]{4,}", co)
    for w in core_words:
        if w.lower() in haystack:
            return True, f"partial:{w}"
    return False, ""

plans_strict = []
plans_partial = []
ambig = []
for r in target:
    co = r.get("company") or ""
    dt = (r.get("responseDate") or "")[:10]
    rk = r.get("registryKey")
    crs = r.get("course") or ""
    strict_cands = {}
    partial_cands = {}
    for c in all_contracts:
        co_m, ctype = co_match(co, c)
        if not co_m: continue
        if not date_in(dt, c["schedule"], window=14): continue
        target_dict = strict_cands if ctype == "strict" else partial_cands
        target_dict.setdefault(c["name"], []).append((c, ctype))

    # v24-30: ambig 정리 — 일정이 응답일과 가까운 strict cand 우선
    def closest_strict(cands_dict):
        """strict cand 중 response_date 매칭이 가장 정확한 1명 선택"""
        if len(cands_dict) <= 1: return None
        # 각 cand의 일정에 정확히 dt 포함하는 contract 찾기
        best = []
        for name, contracts in cands_dict.items():
            for c, _ in contracts:
                # date_in window=0 strict
                if date_in(dt, c["schedule"], window=0):
                    best.append((name, c))
        if len(best) == 1:
            return best[0]
        return None

    if len(strict_cands) == 1:
        name = list(strict_cands.keys())[0]
        c, ctype = strict_cands[name][0]
        plans_strict.append({"rk": rk, "name": name, "co": co, "crs": crs, "dt": dt, "ev": f"{c['sheet'][:5]}_r{c['row']}_{ctype}"})
    elif len(strict_cands) == 0 and len(partial_cands) == 1:
        name = list(partial_cands.keys())[0]
        c, ctype = partial_cands[name][0]
        plans_partial.append({"rk": rk, "name": name, "co": co, "crs": crs, "dt": dt, "ev": f"{c['sheet'][:5]}_r{c['row']}_{ctype}"})
    elif strict_cands:
        # strict ambig — closest_strict로 좁히기
        chosen = closest_strict(strict_cands)
        if chosen:
            name, c = chosen
            plans_strict.append({"rk": rk, "name": name, "co": co, "crs": crs, "dt": dt, "ev": f"{c['sheet'][:5]}_r{c['row']}_strict_closest"})
        else:
            combo = {**strict_cands, **partial_cands}
            ambig.append({"co": co, "dt": dt, "names": list(combo.keys())})
    elif partial_cands:
        combo = {**strict_cands, **partial_cands}
        ambig.append({"co": co, "dt": dt, "names": list(combo.keys())})

print(f"\nstrict 매칭: {len(plans_strict)}")
print(f"partial 매칭: {len(plans_partial)}")
print(f"ambiguous: {len(ambig)}")
print()
print("=== strict plans ===")
for p in plans_strict:
    print(f"  {p['co'][:14]:14s} | {p['dt']} | {p['crs'][:25]:25s} -> {p['name']:6} [{p['ev']}]")
print()
print("=== partial plans ===")
for p in plans_partial:
    print(f"  {p['co'][:14]:14s} | {p['dt']} | {p['crs'][:25]:25s} -> {p['name']:6} [{p['ev']}]")
print()
print("=== ambiguous (candidates 다중) ===")
for a in ambig[:30]:
    print(f"  {a['co'][:14]:14s} | {a['dt']} | {a['names']}")

# generate curl commands
with open(r"C:\Users\Day1_김민선\Downloads\instructor_db\curl_cmds.sh", "w", encoding="utf-8") as f:
    f.write("#!/bin/bash\nset +e\n")
    for p in plans_strict + plans_partial:
        rk_enc = urllib.parse.quote(p["rk"], safe="")
        name_enc = urllib.parse.quote(p["name"], safe="")
        ev_enc = urllib.parse.quote(p["ev"], safe="")
        f.write(f'echo "→ {p["co"][:14]} → {p["name"]}"\n')
        f.write(f'curl -sS -X POST -H "x-cron-secret: $CRON_SECRET" "https://instructor-dashboard.skillflo.app/api/admin/manual-resolve-registry?registry_key={rk_enc}&instructor_name={name_enc}&basis={ev_enc}" --max-time 30\n')
        f.write("echo\n")
