import openpyxl, json, io, sys, re, urllib.request, urllib.parse, ssl, os, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

SECRET = os.environ["CRON_SECRET"]
ctx = ssl.create_default_context()

# 1. contract data load
xlsx_path = r"C:\Users\Day1_김민선\Downloads\instructor_db\contract.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
all_contracts = []
sheets_config = [
    ("강사_일반계약요청 (사용XXXXX)", 9, 13, 14, 7),
    ("강사_일반계약요청의 사본(달리수정중)", 10, 15, 16, 8),
    ("강사_일반계약요청_삼성전자", 8, 12, 13, 6),
]
for sname, name_col, sched_col, loc_col, link_col in sheets_config:
    ws = wb[sname]
    for r in range(3, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col).value
        sched = ws.cell(row=r, column=sched_col).value
        loc = ws.cell(row=r, column=loc_col).value
        link = ws.cell(row=r, column=link_col).value
        if name and isinstance(name, str) and 2 <= len(name) <= 8:
            all_contracts.append({
                "name": name.strip(),
                "schedule": str(sched or ""),
                "location": str(loc or ""),
                "link": str(link or ""),
                "sheet": sname,
                "row": r,
            })
print(f"contract rows: {len(all_contracts)}")

# 2. pending registry load
all_rows = []
for off in [0,100,200,300]:
    try:
        with open(rf"C:\Users\Day1_김민선\Downloads\instructor_db\p{off}.json", "rb") as f:
            all_rows.extend(json.loads(f.read()).get("rows", []))
    except: pass

PUBLIC_KW = ["대학교","학교","진흥원","회의소","공단","연구원","공사","수력원자력","교육원","벤처캐피탈","과기대","과학기술원"]
def is_public(co):
    if not co: return False
    return any(k in co for k in PUBLIC_KW)

target_rows = [r for r in all_rows if not is_public(r.get("company")) and r.get("company") and r.get("responseDate")]
print(f"분석 대상: {len(target_rows)}")

def date_in(target_dt, sched_str):
    if not target_dt or not sched_str: return False
    try:
        y, m, d = target_dt.split("-")
        patterns = [
            f"{y}[.\\-/년 ]+{int(m):02d}[.\\-/월 ]+{int(d):02d}",
            f"{y}[.\\-/년 ]+{int(m)}[.\\-/월 ]+{int(d)}",
            f"{y}\\.\\s*{int(m):02d}\\.\\s*{int(d):02d}",
        ]
        for p in patterns:
            if re.search(p, sched_str): return True
    except: pass
    return False

def co_in(co, contract):
    co_clean = co.replace(" ", "").replace("-", "").lower()
    haystack = (contract["location"] + " " + contract["schedule"] + " " + contract["link"]).lower().replace(" ","")
    if not co_clean: return False
    return co_clean in haystack

plans = []
skipped_ambig = []
for r in target_rows:
    co = r.get("company") or ""
    dt = (r.get("responseDate") or "")[:10]
    rk = r.get("registryKey")
    crs = r.get("course") or ""
    candidates = {}
    for c in all_contracts:
        if co_in(co, c) and date_in(dt, c["schedule"]):
            candidates.setdefault(c["name"], []).append(c)
    if len(candidates) == 1:
        name = list(candidates.keys())[0]
        c = candidates[name][0]
        plans.append({
            "registry_key": rk,
            "instructor_name": name,
            "company": co,
            "course": crs[:40],
            "response_date": dt,
            "evidence": f"{c['sheet'][:5]}/r{c['row']}",
        })
    elif len(candidates) >= 2:
        skipped_ambig.append({"reg": rk, "co": co, "dt": dt, "names": list(candidates.keys())})

print(f"\n=== 단일 후보 (apply 대상): {len(plans)} ===")
for p in plans:
    print(f"  {p['company'][:14]:14s} | {p['response_date']} | {p['course'][:25]:25s} -> {p['instructor_name']} [{p['evidence']}]")
print(f"\n=== ambiguous (다중): {len(skipped_ambig)} ===")
for s in skipped_ambig[:15]:
    print(f"  {s['co'][:14]:14s} | {s['dt']} | candidates: {s['names']}")

# apply
applied = 0
errors = []
for p in plans:
    url = f"https://instructor-dashboard.skillflo.app/api/admin/manual-resolve-registry?registry_key={urllib.parse.quote(p['registry_key'])}&instructor_name={urllib.parse.quote(p['instructor_name'])}&basis={urllib.parse.quote(p['evidence'])}"
    req = urllib.request.Request(url, method="POST", headers={"x-cron-secret": SECRET})
    try:
        with urllib.request.urlopen(req, timeout=30, context=ctx) as r:
            res = json.loads(r.read())
        if res.get("ok"):
            applied += 1
        else:
            errors.append((p, res.get("error")))
    except urllib.error.HTTPError as e:
        try:
            body = json.loads(e.read())
            errors.append((p, body.get("error", str(e))))
        except:
            errors.append((p, str(e)))
    except Exception as e:
        errors.append((p, str(e)))
    time.sleep(0.2)

print(f"\nApplied: {applied}/{len(plans)}")
if errors:
    print("Errors:")
    for p, err in errors[:10]:
        print(f"  {p['company'][:12]:12s} → {p['instructor_name']}: {err}")
